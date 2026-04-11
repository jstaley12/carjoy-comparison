import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import warnings
warnings.filterwarnings("ignore")

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="CarJoy | Comparison Tool",
    page_icon="🚗",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── Password gate ─────────────────────────────────────────────────────────────
PASSWORD = "carjoy2025"

def check_password():
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    if st.session_state.authenticated:
        return True

    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@700;900&family=DM+Sans:wght@300;400;500&display=swap');
    html, body, [data-testid="stAppViewContainer"] {
        background: #f7f4ef;
    }
    .login-wrap {
        display: flex; flex-direction: column; align-items: center;
        justify-content: center; min-height: 80vh; gap: 0;
    }
    .login-logo {
        font-family: 'Playfair Display', serif; font-size: 3.2rem; font-weight: 900;
        color: #1a1a1a; letter-spacing: -1px; margin-bottom: 4px;
    }
    .login-logo span { color: #d4900a; }
    .login-sub {
        font-family: 'DM Sans', sans-serif; font-size: 0.75rem;
        color: #999; letter-spacing: 4px; text-transform: uppercase;
        margin-bottom: 40px;
    }
    </style>
    <div class="login-wrap">
        <div class="login-logo">Car<span>Joy</span></div>
        <div class="login-sub">Pricing Intelligence</div>
    </div>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 1.2, 1])
    with col2:
        pwd = st.text_input("Password", type="password", label_visibility="collapsed",
                            placeholder="Enter password")
        if st.button("Sign In", use_container_width=True):
            if pwd == PASSWORD:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("Incorrect password.")
    return False

# ── Styling ───────────────────────────────────────────────────────────────────
MAIN_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@600;700;800&family=DM+Sans:ital,wght@0,300;0,400;0,500;1,300&display=swap');

html, body, [data-testid="stAppViewContainer"], [data-testid="stHeader"] {
    background: #0a0a0f !important;
    color: #e8e6e0;
}
[data-testid="stSidebar"] { background: #0f0f18 !important; }

h1,h2,h3 { font-family: 'Syne', sans-serif !important; }
p, li, span, div, label { font-family: 'DM Sans', sans-serif !important; }

.cj-header {
    font-family: 'Syne', sans-serif;
    font-size: 2.4rem; font-weight: 800;
    color: #fff; letter-spacing: -1.5px;
    margin-bottom: 2px;
}
.cj-header span { color: #f0b429; }
.cj-sub {
    font-family: 'DM Sans', sans-serif;
    font-size: 0.78rem; color: #666;
    letter-spacing: 3px; text-transform: uppercase;
    margin-bottom: 32px;
}
.upload-box {
    background: #16162a;
    border: 1px solid #2a2a40;
    border-radius: 12px;
    padding: 24px 20px 20px;
    margin-bottom: 16px;
}
.upload-label {
    font-family: 'Syne', sans-serif;
    font-size: 0.75rem; font-weight: 700;
    color: #f0b429; letter-spacing: 2px;
    text-transform: uppercase; margin-bottom: 8px;
}
.stat-card {
    background: #16162a; border: 1px solid #2a2a40;
    border-radius: 10px; padding: 16px 20px;
    text-align: center;
}
.stat-val {
    font-family: 'Syne', sans-serif;
    font-size: 1.8rem; font-weight: 800; color: #ffffff;
}
.stat-val.up { color: #4fd17a; }
.stat-val.down { color: #ff6b6b; }
.stat-label {
    font-family: 'DM Sans', sans-serif;
    font-size: 0.72rem; color: #888;
    letter-spacing: 2px; text-transform: uppercase; margin-top: 4px;
}
.issue-card {
    background: #1a1208; border: 1px solid #3d2c00;
    border-radius: 10px; padding: 16px 20px; margin-bottom: 12px;
}
.issue-title {
    font-family: 'Syne', sans-serif;
    font-size: 0.85rem; font-weight: 700; color: #f0b429;
    margin-bottom: 6px;
}
.issue-row { font-size: 0.82rem; color: #bbb; margin-bottom: 3px; }
.issue-row b { color: #fff; }
.section-head {
    font-family: 'Syne', sans-serif; font-size: 1rem;
    font-weight: 700; color: #ffffff; letter-spacing: 1px;
    text-transform: uppercase; border-bottom: 1px solid #2a2a40;
    padding-bottom: 8px; margin: 28px 0 14px;
}
/* Brighter tab styling */
[data-testid="stTabs"] [data-baseweb="tab-list"] {
    background: #16162a !important;
    border-radius: 8px !important;
    padding: 4px !important;
    gap: 2px !important;
}
[data-testid="stTabs"] [data-baseweb="tab"] {
    color: #cccccc !important;
    font-family: 'Syne', sans-serif !important;
    font-weight: 600 !important;
    font-size: 0.8rem !important;
    background: transparent !important;
    border-radius: 6px !important;
    padding: 6px 14px !important;
}
[data-testid="stTabs"] [aria-selected="true"] {
    background: #f0b429 !important;
    color: #0a0a0f !important;
}
[data-testid="stTabs"] [data-baseweb="tab"]:hover {
    color: #f0b429 !important;
    background: #1e1e35 !important;
}
.stButton > button {
    background: #f0b429 !important; color: #0a0a0f !important;
    font-family: 'Syne', sans-serif !important; font-weight: 700 !important;
    letter-spacing: 1px !important; border: none !important;
    border-radius: 8px !important; padding: 12px 32px !important;
    font-size: 0.9rem !important;
}
.stButton > button:hover { background: #ffc94d !important; }
.stDownloadButton > button {
    background: #16162a !important; color: #f0b429 !important;
    border: 1px solid #f0b429 !important;
    font-family: 'Syne', sans-serif !important; font-weight: 700 !important;
}
.stDownloadButton > button:hover {
    background: #f0b429 !important; color: #0a0a0f !important;
}
[data-testid="stFileUploader"] {
    background: #fff !important;
    border-radius: 8px !important;
}
[data-testid="stFileUploaderDropzone"] {
    background: #f7f4ef !important;
    border: 1.5px dashed #e8e2d9 !important;
}
[data-testid="stDataFrame"] {
    border: 1.5px solid #e8e2d9 !important;
    border-radius: 8px !important;
}
</style>
"""

# ── Core comparison logic ─────────────────────────────────────────────────────

CATEGORY_LABELS = {
    'small suv','mid size suv','midsize suv','large suvs','large suv',
    'pick up trucks','pickup trucks','ev & hybrid','ev&hybrid',
    'mini- van','minivan','mini-van','sedans','trucks','luxury',
    'sports cars','coupes','convertibles','hatchbacks','wagons'
}

def norm(s):
    if s is None: return ''
    if isinstance(s, float) and np.isnan(s): return ''
    try:
        if pd.isna(s): return ''
    except (ValueError, TypeError):
        pass
    return str(s).strip().lower()

def read_hot_sheet(uploaded_file):
    raw = pd.read_excel(uploaded_file, sheet_name="HOT SHEET", header=None)
    header_row = 1
    for i, row in raw.iterrows():
        row_vals = [norm(v) for v in row.values]
        if any('make' in v for v in row_vals):
            header_row = i
            break
    df = pd.read_excel(uploaded_file, sheet_name="HOT SHEET", header=header_row)
    df.columns = [str(c).strip().lower() for c in df.columns]

    def find_col(keywords, exclude=None):
        for c in df.columns:
            if exclude and any(e in c for e in exclude):
                continue
            if any(k in c for k in keywords):
                return c
        return None

    year_col    = find_col(['year'])
    make_col    = find_col(['make'])
    model_col   = find_col(['model'], exclude=['trim'])
    trim_col    = find_col(['trim'])
    msrp_col    = find_col(['msrp'])
    payment_col = find_col(['monthly payment', 'payment'])

    out = pd.DataFrame()
    out['year']    = df[year_col].copy()    if year_col    else np.nan
    out['make']    = df[make_col].copy()    if make_col    else np.nan
    out['model']   = df[model_col].copy()   if model_col   else np.nan
    out['trim']    = df[trim_col].copy()    if trim_col    else np.nan
    out['msrp']    = pd.to_numeric(df[msrp_col],    errors='coerce') if msrp_col    else np.nan
    out['payment'] = pd.to_numeric(df[payment_col], errors='coerce') if payment_col else np.nan
    out['year']    = pd.to_numeric(out['year'], errors='coerce')

    out = out[~out['make'].apply(norm).isin(CATEGORY_LABELS)]
    out = out[~(out['make'].isna() & out['model'].isna() & out['trim'].isna())]
    out = out.reset_index(drop=True)
    return out

def classify_row(row, field='make'):
    val = str(row.get(field, '') or '')
    # Date in make
    if field == 'make':
        import re
        if re.match(r'\d{4}-\d{2}-\d{2}', val.strip()):
            return 'Data Error', 'Vehicle Make contains a date instead of a manufacturer name'
        if val.strip().replace('.','').replace('-','').isdigit() and val.strip():
            return 'Data Error', 'Vehicle Make contains a numeric value'
    # MSRP checks
    msrp = row.get('msrp', None)
    payment = row.get('payment', None)
    if pd.notna(msrp):
        try:
            m = float(msrp)
            if m < 5000:
                return 'Data Error', f'MSRP of ${m:,.0f} is below $5,000 — likely a data entry error'
            if pd.notna(payment):
                p = float(payment)
                if p > 0 and abs(m - p) / max(p, 1) < 0.05:
                    return 'Data Error', f'MSRP (${m:,.0f}) appears to equal Monthly Payment — possible column swap'
        except:
            pass
    return None, None

def make_key(row):
    return (norm(row.get('year','')), norm(row.get('make','')),
            norm(row.get('model','')), norm(row.get('trim','')))

def run_comparison(prev_df, curr_df):
    results = {
        'msrp_inc': [], 'msrp_dec': [],
        'pay_inc': [], 'pay_dec': [],
        'added': [], 'removed': [],
        'unmatched': [], 'data_issues': []
    }

    # Build lookup dicts — list of rows per key
    def build_lookup(df):
        d = {}
        for _, row in df.iterrows():
            k = make_key(row)
            if k not in d:
                d[k] = []
            d[k].append(row)
        return d

    prev_lookup = build_lookup(prev_df)
    curr_lookup = build_lookup(curr_df)

    all_keys = set(prev_lookup.keys()) | set(curr_lookup.keys())

    processed_prev = set()
    processed_curr = set()

    for key in all_keys:
        year, make, model, trim = key
        in_prev = key in prev_lookup
        in_curr = key in curr_lookup

        # Skip if key fields are all empty
        if not make and not model:
            continue

        prev_rows = prev_lookup.get(key, [])
        curr_rows = curr_lookup.get(key, [])

        # ── Missing key field handling ──
        missing_fields = []
        if not year: missing_fields.append('Year')
        if not make: missing_fields.append('Make')
        if not model: missing_fields.append('Model')

        # ── Data error checks ──
        def get_errors(rows, file_label):
            errs = []
            for r in rows:
                err_type, err_msg = classify_row(dict(r), 'make')
                if err_type:
                    errs.append((r, file_label, err_msg))
                else:
                    err_type2, err_msg2 = classify_row(dict(r), 'msrp')
                    if err_type2:
                        errs.append((r, file_label, err_msg2))
            return errs

        prev_errors = get_errors(prev_rows, 'Previous Month')
        curr_errors = get_errors(curr_rows, 'Current Month')

        has_error = bool(prev_errors or curr_errors)
        match_quality = 'Data Error' if has_error else (
            'Incomplete Match' if missing_fields else 'Exact Match'
        )

        # ── ADDED (only in current) ──
        if not in_prev and in_curr:
            for r in curr_rows:
                mq = 'Data Error' if get_errors([r], 'Current Month') else (
                    'Incomplete Match' if missing_fields else 'Unmatched'
                )
                row_out = {
                    'Year': r.get('year'), 'Make': r.get('make'), 'Model': r.get('model'),
                    'Trim': r.get('trim'), 'Old MSRP': np.nan, 'New MSRP': r.get('msrp'),
                    'MSRP Change': np.nan, 'Old Monthly Payment': np.nan,
                    'New Monthly Payment': r.get('payment'), 'Monthly Payment Change': np.nan,
                    'Match Quality': mq
                }
                results['added'].append(row_out)
                if mq == 'Incomplete Match':
                    results['unmatched'].append(row_out)
                # Log data errors
                for (er, fl, em) in get_errors([r], 'Current Month'):
                    results['data_issues'].append({
                        'Vehicle': f"{r.get('year','')} {r.get('make','')} {r.get('model','')} {r.get('trim','')}".strip(),
                        'File': fl, 'Issue': em,
                        'Impact': 'Vehicle appears in Added tab with Data Error flag',
                        'Fix': 'Correct the value in the source HOT SHEET'
                    })
            continue

        # ── REMOVED (only in previous) ──
        if in_prev and not in_curr:
            for r in prev_rows:
                mq = 'Data Error' if get_errors([r], 'Previous Month') else (
                    'Incomplete Match' if missing_fields else 'Unmatched'
                )
                row_out = {
                    'Year': r.get('year'), 'Make': r.get('make'), 'Model': r.get('model'),
                    'Trim': r.get('trim'), 'Old MSRP': r.get('msrp'), 'New MSRP': np.nan,
                    'MSRP Change': np.nan, 'Old Monthly Payment': r.get('payment'),
                    'New Monthly Payment': np.nan, 'Monthly Payment Change': np.nan,
                    'Match Quality': mq
                }
                results['removed'].append(row_out)
                if mq == 'Incomplete Match':
                    results['unmatched'].append(row_out)
                for (er, fl, em) in get_errors([r], 'Previous Month'):
                    results['data_issues'].append({
                        'Vehicle': f"{r.get('year','')} {r.get('make','')} {r.get('model','')} {r.get('trim','')}".strip(),
                        'File': fl, 'Issue': em,
                        'Impact': 'Vehicle appears in Removed tab with Data Error flag',
                        'Fix': 'Correct the value in the source HOT SHEET'
                    })
            continue

        # ── MATCHED — in both ──
        # For duplicate trims: generate all prev->curr pairs where MSRP changed
        # This correctly captures each distinct price point transition
        if len(prev_rows) == 1 and len(curr_rows) == 1:
            pairs = [(prev_rows[0], curr_rows[0])]
        else:
            # Cross-product: match each prev MSRP to closest curr MSRP
            # Only keep pairs where values actually differ (real changes)
            pairs = []
            used_curr = set()
            # Sort both by MSRP to align price tiers
            prev_sorted = sorted(prev_rows, key=lambda r: float(r.get('msrp') or 0))
            curr_sorted = sorted(curr_rows, key=lambda r: float(r.get('msrp') or 0))
            for pr in prev_sorted:
                best_ci, best_diff = None, float('inf')
                pm = pr.get('msrp')
                for ci, cr in enumerate(curr_sorted):
                    cm = cr.get('msrp')
                    if pd.notna(pm) and pd.notna(cm):
                        diff = abs(float(pm) - float(cm))
                    else:
                        diff = 999999
                    if diff < best_diff:
                        best_diff, best_ci = diff, ci
                if best_ci is not None:
                    pairs.append((pr, curr_sorted[best_ci]))
                    used_curr.add(best_ci)
            # Unpaired curr rows → flag as unmatched
            for ci, cr in enumerate(curr_sorted):
                if ci not in used_curr:
                    row_out = {
                        'Year': cr.get('year'), 'Make': cr.get('make'),
                        'Model': cr.get('model'), 'Trim': cr.get('trim'),
                        'Old MSRP': np.nan, 'New MSRP': cr.get('msrp'),
                        'MSRP Change': np.nan, 'Old Monthly Payment': np.nan,
                        'New Monthly Payment': cr.get('payment'),
                        'Monthly Payment Change': np.nan, 'Match Quality': 'Possible Duplicate'
                    }
                    results['unmatched'].append(row_out)

        for (pr, cr) in pairs:
            # Determine match quality
            mq = match_quality
            if len(prev_rows) > 1 or len(curr_rows) > 1:
                mq = 'Possible Duplicate' if not has_error else 'Data Error'
            elif missing_fields:
                mq = 'Incomplete Match'

            # Compute changes
            old_msrp = pr.get('msrp')
            new_msrp = cr.get('msrp')
            old_pay  = pr.get('payment')
            new_pay  = cr.get('payment')

            msrp_chg = np.nan
            pay_chg  = np.nan
            try:
                if pd.notna(old_msrp) and pd.notna(new_msrp):
                    msrp_chg = float(new_msrp) - float(old_msrp)
            except: pass
            try:
                if pd.notna(old_pay) and pd.notna(new_pay):
                    pay_chg = float(new_pay) - float(old_pay)
            except: pass

            row_out = {
                'Year': cr.get('year'), 'Make': cr.get('make'),
                'Model': cr.get('model'), 'Trim': cr.get('trim'),
                'Old MSRP': old_msrp, 'New MSRP': new_msrp,
                'MSRP Change': msrp_chg,
                'Old Monthly Payment': old_pay, 'New Monthly Payment': new_pay,
                'Monthly Payment Change': pay_chg,
                'Match Quality': mq
            }

            # Log data issues
            for (er, fl, em) in (prev_errors + curr_errors):
                vehicle_str = f"{cr.get('year','')} {cr.get('make','')} {cr.get('model','')} {cr.get('trim','')}".strip()
                impact = ''
                if pd.notna(msrp_chg):
                    impact = f"Caused a reported MSRP change of ${msrp_chg:+,.0f} which may be invalid"
                elif pd.notna(pay_chg):
                    impact = f"Caused a reported payment change of ${pay_chg:+,.0f} which may be invalid"
                else:
                    impact = "Row flagged — verify manually before using in reports"
                results['data_issues'].append({
                    'Vehicle': vehicle_str, 'File': fl, 'Issue': em,
                    'Impact': impact,
                    'Fix': 'Correct the value in the source HOT SHEET file'
                })

            # Incomplete match → also goes to unmatched
            if mq in ('Incomplete Match', 'Possible Duplicate'):
                results['unmatched'].append(row_out)

            # Route to change tabs
            if pd.notna(msrp_chg):
                if msrp_chg > 0:
                    results['msrp_inc'].append(row_out)
                elif msrp_chg < 0:
                    results['msrp_dec'].append(row_out)

            if pd.notna(pay_chg):
                if pay_chg > 0:
                    results['pay_inc'].append(row_out)
                elif pay_chg < 0:
                    results['pay_dec'].append(row_out)

    return results

# ── Excel builder ─────────────────────────────────────────────────────────────

COLS = ['Year','Make','Model','Trim','Old MSRP','New MSRP','MSRP Change',
        'Old Monthly Payment','New Monthly Payment','Monthly Payment Change','Match Quality']

def style_header(ws, row=1):
    header_fill = PatternFill("solid", fgColor="0A0A0F")
    accent_fill  = PatternFill("solid", fgColor="F0B429")
    for cell in ws[row]:
        cell.fill = accent_fill
        cell.font = Font(bold=True, color="0A0A0F", size=10,
                         name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center")

def style_data(ws, start_row=2):
    light = PatternFill("solid", fgColor="12121C")
    dark  = PatternFill("solid", fgColor="0E0E1A")
    for i, row in enumerate(ws.iter_rows(min_row=start_row, max_row=ws.max_row)):
        fill = light if i % 2 == 0 else dark
        for cell in row:
            cell.fill = fill
            cell.font = Font(color="E8E6E0", size=9, name="Calibri")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            # Color MSRP/Payment change columns
            col_name = ws.cell(row=1, column=cell.column).value
            if col_name in ('MSRP Change', 'Monthly Payment Change'):
                try:
                    v = float(cell.value)
                    cell.font = Font(
                        color="5EBD6E" if v > 0 else ("E05C5C" if v < 0 else "888888"),
                        bold=True, size=9, name="Calibri"
                    )
                except: pass
            if col_name == 'Match Quality':
                mq_colors = {
                    'Data Error': 'E05C5C',
                    'Incomplete Match': 'F0B429',
                    'Possible Duplicate': 'C47D2B',
                    'Exact Match': '5EBD6E',
                    'Close Match': '5EBD6E',
                    'Unmatched': '888888',
                }
                color = mq_colors.get(str(cell.value), 'E8E6E0')
                cell.font = Font(color=color, bold=True, size=9, name="Calibri")

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except: pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 40)

def write_tab(wb, title, data, cols=COLS):
    ws = wb.create_sheet(title=title)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "F0B429"
    bg = PatternFill("solid", fgColor="0A0A0F")
    # Write header
    for ci, col in enumerate(cols, 1):
        ws.cell(row=1, column=ci, value=col)
    # Write data
    df = pd.DataFrame(data, columns=cols) if data else pd.DataFrame(columns=cols)
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci)
            if pd.isna(val) if not isinstance(val, str) else False:
                cell.value = None
            else:
                try:
                    cell.value = round(float(val), 2) if isinstance(val, float) else val
                except:
                    cell.value = val
    style_header(ws)
    style_data(ws)
    auto_width(ws)
    ws.row_dimensions[1].height = 22
    return ws, df

def build_excel(results):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    tab_map = [
        ("MSRP Increased",           results['msrp_inc']),
        ("MSRP Decreased",           results['msrp_dec']),
        ("Payment Increased",        results['pay_inc']),
        ("Payment Decreased",        results['pay_dec']),
        ("Vehicles Added",           results['added']),
        ("Vehicles Removed",         results['removed']),
        ("Unmatched Review",         results['unmatched'][:25]),
    ]

    tab_dfs = {}
    for title, data in tab_map:
        ws, df = write_tab(wb, title, data)
        tab_dfs[title] = df

    # ── Summary Dashboard ──────────────────────────────────────────────────
    ws = wb.create_sheet(title="Summary Dashboard")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "F0B429"
    bg_fill   = PatternFill("solid", fgColor="0A0A0F")
    card_fill = PatternFill("solid", fgColor="12121C")
    accent    = PatternFill("solid", fgColor="F0B429")
    head_font = Font(name="Calibri", bold=True, color="F0B429", size=11)
    val_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    lbl_font  = Font(name="Calibri", color="666666", size=8)
    body_font = Font(name="Calibri", color="E8E6E0", size=9)
    up_font   = Font(name="Calibri", bold=True, color="5EBD6E", size=13)
    dn_font   = Font(name="Calibri", bold=True, color="E05C5C", size=13)

    def set_bg(ws):
        for row in ws.iter_rows(min_row=1, max_row=200, min_col=1, max_col=20):
            for cell in row:
                cell.fill = bg_fill

    set_bg(ws)

    r = 1
    # Title
    ws.merge_cells(f"B{r}:J{r}")
    tc = ws[f"B{r}"]
    tc.value = "CARJOY  |  MONTHLY PRICING COMPARISON"
    tc.font = Font(name="Calibri", bold=True, color="F0B429", size=16)
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 32
    r += 1

    ws.merge_cells(f"B{r}:J{r}")
    sc = ws[f"B{r}"]
    sc.value = "SUMMARY DASHBOARD"
    sc.font = Font(name="Calibri", color="444444", size=8)
    sc.alignment = Alignment(horizontal="left")
    r += 2

    # ── Count cards ──
    def count_card(ws, row, col, label, value, font=None):
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row+1, end_column=col+1)
        c = ws.cell(row=row, column=col)
        c.value = value
        c.font = font or val_font
        c.fill = card_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 26
        lc = ws.cell(row=row+2, column=col)
        lc.value = label
        lc.font = lbl_font
        lc.fill = card_fill
        lc.alignment = Alignment(horizontal="center")
        ws.row_dimensions[row+2].height = 14

    counts = [
        ("MSRP ↑",     len(results['msrp_inc']),  up_font),
        ("MSRP ↓",     len(results['msrp_dec']),  dn_font),
        ("PMT ↑",      len(results['pay_inc']),   up_font),
        ("PMT ↓",      len(results['pay_dec']),   dn_font),
        ("ADDED",      len(results['added']),      val_font),
        ("REMOVED",    len(results['removed']),    val_font),
        ("DATA ISSUES",len(results['data_issues']),
         Font(name="Calibri", bold=True, color="E05C5C", size=13)),
    ]
    col_start = 2
    for label, value, font in counts:
        count_card(ws, r, col_start, label, value, font)
        col_start += 3
    r += 5

    # ── Averages ──
    def safe_avg(lst, field):
        vals = [float(x[field]) for x in lst
                if x.get(field) is not None and not (
                    isinstance(x[field], float) and np.isnan(x[field]))]
        return round(sum(vals)/len(vals), 2) if vals else 0

    def safe_max(lst, field):
        vals = [float(x[field]) for x in lst
                if x.get(field) is not None and not (
                    isinstance(x[field], float) and np.isnan(x[field]))]
        return max(vals) if vals else 0

    def safe_min(lst, field):
        vals = [float(x[field]) for x in lst
                if x.get(field) is not None and not (
                    isinstance(x[field], float) and np.isnan(x[field]))]
        return min(vals) if vals else 0

    avg_data = [
        ("Avg MSRP Increase",   safe_avg(results['msrp_inc'], 'MSRP Change'),   "+$"),
        ("Avg MSRP Decrease",   safe_avg(results['msrp_dec'], 'MSRP Change'),   "$"),
        ("Avg Payment Increase",safe_avg(results['pay_inc'],  'Monthly Payment Change'), "+$"),
        ("Avg Payment Decrease",safe_avg(results['pay_dec'],  'Monthly Payment Change'), "$"),
        ("Largest MSRP Increase",  safe_max(results['msrp_inc'], 'MSRP Change'), "+$"),
        ("Largest MSRP Decrease",  safe_min(results['msrp_dec'], 'MSRP Change'), "$"),
    ]

    ws.cell(row=r, column=2).value = "AVERAGES & EXTREMES"
    ws.cell(row=r, column=2).font = head_font
    r += 1

    for i, (label, value, prefix) in enumerate(avg_data):
        col = 2 + (i % 3) * 4
        if i > 0 and i % 3 == 0:
            r += 3
        lc = ws.cell(row=r, column=col)
        lc.value = label
        lc.font = lbl_font
        lc.fill = card_fill
        vc = ws.cell(row=r+1, column=col)
        vc.value = f"{prefix}{abs(value):,.2f}"
        vc.font = Font(name="Calibri", bold=True,
                       color="5EBD6E" if value >= 0 else "E05C5C", size=11)
        vc.fill = card_fill
    r += 5

    # ── Top 10 Payment Increases ──
    def write_top10(ws, row, col, title, data, field, ascending=False):
        df = pd.DataFrame(data)
        if df.empty or field not in df.columns:
            return row
        df[field] = pd.to_numeric(df[field], errors='coerce')
        df = df.dropna(subset=[field])
        df = df.sort_values(field, ascending=ascending).head(10)

        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+5)
        tc = ws.cell(row=row, column=col)
        tc.value = title
        tc.font = head_font
        row += 1

        headers = ['Year','Make','Model','Trim','Old Pmt','New Pmt','Change']
        for ci, h in enumerate(headers):
            c = ws.cell(row=row, column=col+ci)
            c.value = h
            c.font = Font(name="Calibri", bold=True, color="0A0A0F", size=9)
            c.fill = accent
            c.alignment = Alignment(horizontal="center")
        row += 1

        light = PatternFill("solid", fgColor="12121C")
        dark  = PatternFill("solid", fgColor="0E0E1A")
        for ri, (_, r_data) in enumerate(df.iterrows()):
            fill = light if ri % 2 == 0 else dark
            vals = [
                r_data.get('Year',''), r_data.get('Make',''),
                r_data.get('Model',''), r_data.get('Trim',''),
                r_data.get('Old Monthly Payment',''),
                r_data.get('New Monthly Payment',''),
                r_data.get(field,'')
            ]
            for ci, v in enumerate(vals):
                c = ws.cell(row=row, column=col+ci)
                try:
                    c.value = round(float(v), 2) if isinstance(v, (int,float)) else v
                except:
                    c.value = v
                c.fill = fill
                is_change = ci == 6
                if is_change:
                    try:
                        fv = float(v)
                        c.font = Font(name="Calibri", bold=True,
                                      color="5EBD6E" if fv>0 else "E05C5C", size=9)
                    except:
                        c.font = Font(name="Calibri", color="E8E6E0", size=9)
                else:
                    c.font = Font(name="Calibri", color="E8E6E0", size=9)
            row += 1
        return row + 1

    r = write_top10(ws, r, 2, "TOP 10 PAYMENT INCREASES",
                    results['pay_inc'], 'Monthly Payment Change', ascending=False)
    r = write_top10(ws, r, 2, "TOP 10 PAYMENT DECREASES",
                    results['pay_dec'], 'Monthly Payment Change', ascending=True)

    # ── Top Movers ──
    ws.cell(row=r, column=2).value = "TOP MOVERS"
    ws.cell(row=r, column=2).font = head_font
    r += 1

    for side, data, field, asc, label in [
        ("TOP 5 PAYMENT INCREASES", results['pay_inc'],
         'Monthly Payment Change', False, "+$"),
        ("TOP 5 PAYMENT DECREASES", results['pay_dec'],
         'Monthly Payment Change', True,  "$"),
    ]:
        df = pd.DataFrame(data)
        ws.cell(row=r, column=2).value = side
        ws.cell(row=r, column=2).font = Font(name="Calibri", bold=True,
                                              color="AAAAAA", size=9)
        r += 1
        if not df.empty and field in df.columns:
            df[field] = pd.to_numeric(df[field], errors='coerce')
            df = df.dropna(subset=[field]).sort_values(field, ascending=asc).head(5)
            for _, row_data in df.iterrows():
                vehicle = f"{row_data.get('Year','')} {row_data.get('Make','')} {row_data.get('Model','')} {row_data.get('Trim','')}".strip()
                chg = row_data.get(field, 0)
                try:
                    chg_str = f"{float(chg):+.0f}/mo"
                    color = "5EBD6E" if float(chg) > 0 else "E05C5C"
                except:
                    chg_str = str(chg); color = "AAAAAA"
                ws.cell(row=r, column=2).value = vehicle
                ws.cell(row=r, column=2).font = Font(name="Calibri", color="E8E6E0", size=9)
                ws.cell(row=r, column=5).value = chg_str
                ws.cell(row=r, column=5).font = Font(name="Calibri", bold=True,
                                                      color=color, size=9)
                r += 1
        r += 1

    # ── Business Insights ──
    ws.cell(row=r, column=2).value = "BUSINESS INSIGHTS"
    ws.cell(row=r, column=2).font = head_font
    r += 1

    pay_inc_count = len(results['pay_inc'])
    pay_dec_count = len(results['pay_dec'])
    msrp_inc_count = len(results['msrp_inc'])
    msrp_dec_count = len(results['msrp_dec'])
    avg_pi = safe_avg(results['pay_inc'], 'Monthly Payment Change')
    avg_pd = safe_avg(results['pay_dec'], 'Monthly Payment Change')

    insights = [
        f"Payment movement was {'balanced' if abs(pay_inc_count-pay_dec_count)<=4 else 'skewed'}: "
        f"{pay_inc_count} increases (avg +${avg_pi:.0f}/mo) vs "
        f"{pay_dec_count} decreases (avg ${avg_pd:.0f}/mo).",

        f"MSRP changes were more selective: {msrp_inc_count} increases vs "
        f"{msrp_dec_count} decreases, suggesting targeted pricing adjustments "
        f"rather than broad market movement.",

        f"{len(results['added'])} vehicles were added and {len(results['removed'])} removed, "
        f"indicating {'moderate' if len(results['added'])+len(results['removed']) < 20 else 'significant'} "
        f"inventory turnover this period.",
    ]

    for ins in insights:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=11)
        c = ws.cell(row=r, column=2)
        c.value = f"• {ins}"
        c.font = Font(name="Calibri", color="AAAAAA", size=9, italic=True)
        c.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 18
        r += 1

    r += 1

    # ── Data Quality Issues ──
    ws.cell(row=r, column=2).value = "⚠️  DATA QUALITY ISSUES"
    ws.cell(row=r, column=2).font = Font(name="Calibri", bold=True,
                                          color="E05C5C", size=11)
    r += 1

    if not results['data_issues']:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
        c = ws.cell(row=r, column=2)
        c.value = "✅  No data quality issues detected."
        c.font = Font(name="Calibri", color="5EBD6E", size=9)
        r += 2
    else:
        seen = set()
        for issue in results['data_issues']:
            key = (issue['Vehicle'], issue['Issue'])
            if key in seen: continue
            seen.add(key)
            issue_fill = PatternFill("solid", fgColor="1A1208")

            fields = [
                ("Vehicle", issue['Vehicle']),
                ("File",    issue['File']),
                ("Issue",   issue['Issue']),
                ("Impact",  issue['Impact']),
                ("Fix",     issue['Fix']),
            ]
            for label, value in fields:
                lc = ws.cell(row=r, column=2)
                lc.value = label + ":"
                lc.font = Font(name="Calibri", bold=True, color="F0B429", size=9)
                lc.fill = issue_fill
                ws.merge_cells(start_row=r, start_column=3,
                               end_row=r, end_column=11)
                vc = ws.cell(row=r, column=3)
                vc.value = value
                vc.font = Font(name="Calibri", color="E8E6E0", size=9)
                vc.fill = issue_fill
                vc.alignment = Alignment(wrap_text=True)
                ws.row_dimensions[r].height = 15
                r += 1
            r += 1

    auto_width(ws)
    ws.column_dimensions['A'].width = 3

    # Move Summary Dashboard to front
    wb.move_sheet("Summary Dashboard", offset=-len(wb.sheetnames)+1)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ── Streamlit UI ──────────────────────────────────────────────────────────────

def main():
    if not check_password():
        return

    st.markdown(MAIN_CSS, unsafe_allow_html=True)

    st.markdown("""
    <div class="cj-header">Car<span>Joy</span></div>
    <div class="cj-sub">Monthly Pricing Intelligence</div>
    """, unsafe_allow_html=True)

    col1, col2 = st.columns(2)

    with col1:
        st.markdown('<div class="upload-box"><div class="upload-label">⬅ Previous Month</div>', unsafe_allow_html=True)
        prev_file = st.file_uploader("Previous Month HOT SHEET",
                                      type=["xlsx","xls"],
                                      key="prev",
                                      label_visibility="collapsed")
        st.markdown('</div>', unsafe_allow_html=True)

    with col2:
        st.markdown('<div class="upload-box"><div class="upload-label">Current Month ➡</div>', unsafe_allow_html=True)
        curr_file = st.file_uploader("Current Month HOT SHEET",
                                      type=["xlsx","xls"],
                                      key="curr",
                                      label_visibility="collapsed")
        st.markdown('</div>', unsafe_allow_html=True)

    if prev_file and curr_file:
        if st.button("Run Comparison →", use_container_width=False):
            with st.spinner("Reading files and running comparison..."):
                try:
                    prev_df = read_hot_sheet(prev_file)
                    curr_df = read_hot_sheet(curr_file)
                    results = run_comparison(prev_df, curr_df)
                    excel_file = build_excel(results)
                    st.session_state['results'] = results
                    st.session_state['excel'] = excel_file
                except Exception as e:
                    st.error(f"Error during comparison: {e}")
                    st.exception(e)
                    return

    if 'results' in st.session_state:
        results = st.session_state['results']

        # ── Stat cards ──
        st.markdown('<div class="section-head">Results Overview</div>', unsafe_allow_html=True)
        cols = st.columns(7)
        stats = [
            ("MSRP ↑",      len(results['msrp_inc']), "up"),
            ("MSRP ↓",      len(results['msrp_dec']), "down"),
            ("PMT ↑",       len(results['pay_inc']),  "up"),
            ("PMT ↓",       len(results['pay_dec']),  "down"),
            ("Added",       len(results['added']),    ""),
            ("Removed",     len(results['removed']),  ""),
            ("Data Issues", len(results['data_issues']), "down" if results['data_issues'] else ""),
        ]
        for col, (label, val, cls) in zip(cols, stats):
            with col:
                st.markdown(f"""
                <div class="stat-card">
                    <div class="stat-val {cls}">{val}</div>
                    <div class="stat-label">{label}</div>
                </div>""", unsafe_allow_html=True)

        # ── Data issues callout ──
        if results['data_issues']:
            st.markdown('<div class="section-head">⚠️ Data Quality Issues</div>', unsafe_allow_html=True)
            seen = set()
            for issue in results['data_issues']:
                key = (issue['Vehicle'], issue['Issue'])
                if key in seen: continue
                seen.add(key)
                st.markdown(f"""
                <div class="issue-card">
                    <div class="issue-title">🚨 {issue['Vehicle']}</div>
                    <div class="issue-row"><b>File:</b> {issue['File']}</div>
                    <div class="issue-row"><b>Issue:</b> {issue['Issue']}</div>
                    <div class="issue-row"><b>Impact:</b> {issue['Impact']}</div>
                    <div class="issue-row"><b>Fix:</b> {issue['Fix']}</div>
                </div>""", unsafe_allow_html=True)

        # ── Preview tabs ──
        st.markdown('<div class="section-head">Preview</div>', unsafe_allow_html=True)
        tab_labels = ["MSRP ↑","MSRP ↓","PMT ↑","PMT ↓","Added","Removed","Unmatched"]
        tab_keys   = ['msrp_inc','msrp_dec','pay_inc','pay_dec','added','removed','unmatched']
        tabs = st.tabs(tab_labels)
        for tab, key in zip(tabs, tab_keys):
            with tab:
                data = results[key]
                if data:
                    df = pd.DataFrame(data, columns=COLS)
                    st.dataframe(df, use_container_width=True, height=300,
                                 hide_index=True)
                else:
                    st.markdown("<p style='color:#555;font-size:0.85rem;padding:12px 0'>No records in this category.</p>", unsafe_allow_html=True)

        # ── Download ──
        st.markdown('<div class="section-head">Download</div>', unsafe_allow_html=True)
        st.download_button(
            label="⬇  Download Full Excel Report",
            data=st.session_state['excel'],
            file_name="carjoy_comparison.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=False,
        )

if __name__ == "__main__":
    main()
